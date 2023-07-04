import openpyxl

class readXLUtil:

    def getRowCount(file , sheet):
        book = openpyxl.load_workbook(file)
        sheet = book[sheet]
        return sheet.max_row

    def readXLData(file , sheet , rowno , colno):
        book = openpyxl.load_workbook(file)
        sheet = book[sheet]
        return sheet.cell(row = rowno , column=colno).value

    def WriteXLData(file , sheet , rowno , colno , data):
        book = openpyxl.load_workbook(file)
        sheet = book[sheet]
        sheet.cell(row = rowno , column=colno).value = data
        book.save(file)